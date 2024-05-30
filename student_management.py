import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from openpyxl import Workbook, load_workbook
import sqlite3
from flask import Flask, jsonify
from tkcalendar import DateEntry
from threading import Thread
import re

class StudentManagementSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Модуль учёта студента факультета")
        self.tabControl = ttk.Notebook(root)
        self.tabControl.pack(expand=1, fill="both")

        self.setup_ui()
        self.setup_database()
        self.load_data()
        self.setup_api()

    def setup_ui(self):
        self.create_tabs()
        self.create_widgets()
        self.create_toolbar()

    def setup_database(self):
        self.conn = sqlite3.connect('student_management.db', check_same_thread=False)
        self.conn.execute('PRAGMA foreign_keys = ON')
        self.cursor = self.conn.cursor()
        self.create_tables()

    def create_tabs(self):
        self.tab_students = ttk.Frame(self.tabControl)
        self.tabControl.add(self.tab_students, text='Студенты')

        self.tab_events = ttk.Frame(self.tabControl)
        # Убираем добавление вкладки "События" из видимого интерфейса
        # self.tabControl.add(self.tab_events, text='События')

        self.tab_groups = ttk.Frame(self.tabControl)
        self.tabControl.add(self.tab_groups, text='Группы')

    def create_tables(self):
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS students (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                first_name TEXT, last_name TEXT, middle_name TEXT, birth_date TEXT,
                phone TEXT, email TEXT, address TEXT, group_id INTEGER,
                FOREIGN KEY(group_id) REFERENCES groups(id) ON DELETE SET NULL
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id INTEGER, date TEXT, title TEXT, description TEXT, category TEXT,
                FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS education_periods (
                student_id INTEGER, start_date TEXT, end_date TEXT, group_name TEXT,
                FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS groups (
                id INTEGER PRIMARY KEY AUTOINCREMENT, group_name TEXT
            )
        ''')
        self.conn.commit()

    def create_widgets(self):
        self.create_student_widgets()
        self.create_event_widgets()
        self.create_group_widgets()

    def create_toolbar(self):
        toolbar = tk.Frame(self.root)
        toolbar.pack(side=tk.TOP, fill=tk.X)

        self.add_button = tk.Button(toolbar, text="Добавить", command=self.add_item)
        self.add_button.pack(side=tk.LEFT, padx=5)

        self.edit_button = tk.Button(toolbar, text="Редактировать", command=self.edit_item)
        self.edit_button.pack(side=tk.LEFT, padx=5)

        self.delete_button = tk.Button(toolbar, text="Удалить", command=self.delete_item)
        self.delete_button.pack(side=tk.LEFT, padx=5)

        self.import_excel_button = tk.Button(toolbar, text="Импорт из Excel", command=self.import_from_excel)
        self.import_excel_button.pack(side=tk.LEFT, padx=5)

        self.export_excel_button = tk.Button(toolbar, text="Экспорт в Excel", command=self.export_to_excel)
        self.export_excel_button.pack(side=tk.LEFT, padx=5)

        self.report_button = tk.Button(toolbar, text="Отчёт по группе за период", command=self.generate_group_report)
        self.report_button.pack(side=tk.LEFT, padx=5)

        self.search_entry = tk.Entry(toolbar)
        self.search_entry.pack(side=tk.LEFT, padx=10, pady=5, fill=tk.X)
        self.search_entry.bind('<KeyRelease>', self.search_item)

    def create_student_widgets(self):
        self.tree_students = self.create_treeview(self.tab_students,
                                                  ['ID', 'Фамилия', 'Имя', 'Отчество', 'Дата рождения', 'Телефон',
                                                   'Email', 'Адрес', 'Группа'])
        self.tree_students.bind("<Button-3>", self.show_student_context_menu)

    def create_event_widgets(self):
        self.tree_events = self.create_treeview(self.tab_events,
                                                ['ID события', 'ID студента', 'Дата', 'Название', 'Описание',
                                                 'Категория'])

    def create_group_widgets(self):
        self.tree_groups = self.create_treeview(self.tab_groups, ['ID', 'Группа'])

    def create_treeview(self, parent, columns):
        frame = tk.Frame(parent)
        frame.pack(fill=tk.BOTH, expand=True)

        treeview = ttk.Treeview(frame, columns=columns, show='headings')
        for col in columns:
            treeview.heading(col, text=col)
            treeview.column(col, width=100)

        h_scrollbar = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=treeview.xview)
        v_scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=treeview.yview)
        treeview.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)

        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        treeview.pack(fill=tk.BOTH, expand=True)

        return treeview

    def load_data(self):
        self.groups = []
        self.students = []
        self.events = []

        self.load_table_data(
            "SELECT students.id, last_name, first_name, middle_name, birth_date, phone, email, address, group_name FROM students LEFT JOIN groups ON students.group_id = groups.id",
            self.tree_students, self.students)
        self.load_table_data("SELECT * FROM events", self.tree_events, self.events)
        self.load_table_data("SELECT * FROM groups", self.tree_groups)

        # Загрузка списка групп
        self.cursor.execute("SELECT id, group_name FROM groups")
        self.groups = self.cursor.fetchall()

    def load_table_data(self, query, treeview, data_list=None):
        self.cursor.execute(query)
        data = self.cursor.fetchall()
        self.update_treeview(treeview, data)
        if data_list is not None:
            data_list.clear()
            data_list.extend(data)

    def update_treeview(self, treeview, data):
        treeview.delete(*treeview.get_children())
        for row in data:
            treeview.insert('', 'end', values=row)

    def add_item(self):
        tab = self.tabControl.select()
        if tab == self.tabControl.tabs()[0]:
            self.open_student_window('Добавить студента', None)
        elif tab == self.tabControl.tabs()[1]:
            self.open_event_window('Добавить событие', None)
        elif tab == self.tabControl.tabs()[2]:
            self.open_group_window('Добавить группу', None)

    def edit_item(self):
        tab = self.tabControl.select()
        if tab == self.tabControl.tabs()[0]:
            selected_item = self.tree_students.selection()
            if selected_item:
                student_id = self.tree_students.item(selected_item)['values'][0]
                self.open_student_window('Редактировать студента', self.get_student_by_id(student_id))
            else:
                messagebox.showwarning("Предупреждение", "Пожалуйста, выберите элемент для редактирования.")
        elif tab == self.tabControl.tabs()[1]:
            self.edit_selected_item(self.tree_events, self.open_event_window)
        elif tab == self.tabControl.tabs()[2]:
            self.edit_selected_item(self.tree_groups, self.open_group_window)

    def delete_item(self):
        tab = self.tabControl.select()
        if tab == self.tabControl.tabs()[0]:
            self.delete_selected_item(self.tree_students, "DELETE FROM students WHERE id=?")
        elif tab == self.tabControl.tabs()[1]:
            self.delete_selected_item(self.tree_events, "DELETE FROM events WHERE id=?")
        elif tab == self.tabControl.tabs()[2]:
            self.delete_selected_item(self.tree_groups, "DELETE FROM groups WHERE id=?")

    def search_item(self, event=None):
        query = self.search_entry.get().lower()
        tab = self.tabControl.select()
        if tab == str(self.tab_students):
            self.search_in_treeview(query, self.students, self.tree_students)
        elif tab == str(self.tab_events):
            self.search_in_treeview(query, self.events, self.tree_events)
        elif tab == str(self.tab_groups):
            self.search_in_treeview(query, self.groups, self.tree_groups)

    def open_student_window(self, title, values=None):
        self.open_entry_window(title,
                               ['Фамилия', 'Имя', 'Отчество', 'Дата рождения', 'Телефон', 'Email', 'Адрес', 'Группа'],
                               self.save_student, values)

    def open_event_window(self, title, values=None, student_id=None):
        self.open_entry_window(title, ['Дата', 'Название', 'Описание', 'Категория'], self.save_event, values, student_id)

    def open_group_window(self, title, values=None):
        window = tk.Toplevel(self.root)
        window.title(title)

        tk.Label(window, text="Группа").grid(row=0, column=0, padx=10, pady=5)
        group_name_entry = tk.Entry(window)
        group_name_entry.grid(row=0, column=1, padx=10, pady=5)

        if values:
            group_name_entry.insert(0, values[1])

        tk.Button(window, text="Сохранить",
                  command=lambda: self.save_group(group_name_entry.get(), window, values[0] if values else None)).grid(
            row=1, column=0, columnspan=2, pady=10)

    def open_entry_window(self, title, labels, save_command, values=None, student_id=None):
        window = tk.Toplevel(self.root)
        window.title(title)

        entries = {}
        for i, label_text in enumerate(labels):
            tk.Label(window, text=label_text).grid(row=i, column=0, padx=10, pady=5)
            entry = DateEntry(window, date_pattern='dd.MM.yyyy') if 'Дата' in label_text else tk.Entry(window)
            if label_text == 'Группа':
                entry = ttk.Combobox(window, values=[group[1] for group in self.groups])
            entry.grid(row=i, column=1, padx=10, pady=5)
            entries[label_text] = entry

        if values:
            for label, value in zip(labels, values[1:]):
                if 'Дата' in label:
                    if value:
                        entries[label].set_date(value)
                else:
                    entries[label].insert(0, value)

        tk.Button(window, text="Сохранить",
                  command=lambda: self.validate_and_save(entries, save_command, window, values[0] if values else None, student_id)).grid(
            row=len(labels), column=0, columnspan=2, pady=10)

    def validate_email(self, email):
        email_regex = re.compile(r"[^@]+@[^@]+\.[^@]+")
        return email_regex.match(email) is not None

    def validate_phone(self, phone):
        phone_regex = re.compile(r"^\+?\d{10,15}$")
        return phone_regex.match(phone) is not None

    def validate_and_save(self, entries, save_command, window, record_id=None, student_id=None):
        for label, entry in entries.items():
            if not entry.get().strip():
                messagebox.showerror("Ошибка валидации", f"Пожалуйста, заполните поле '{label}'.")
                return

        if 'Email' in entries and not self.validate_email(entries['Email'].get()):
            messagebox.showerror("Ошибка валидации", "Пожалуйста, введите корректный email.")
            return

        if 'Телефон' in entries and not self.validate_phone(entries['Телефон'].get()):
            messagebox.showerror("Ошибка валидации", "Пожалуйста, введите корректный номер телефона.")
            return

        if save_command == self.save_student:
            save_command(entries, window, record_id)
        else:
            save_command(entries, window, record_id, student_id)

    def save_student(self, entries, window, student_id=None):
        data = {label: entry.get() for label, entry in entries.items()}
        group_id = self.get_group_id(data['Группа'])

        # Проверка на уникальность email и телефона
        if not self.check_unique_email(data['Email'], student_id):
            messagebox.showerror("Ошибка", "Email уже существует.")
            return

        if not self.check_unique_phone(data['Телефон'], student_id):
            messagebox.showerror("Ошибка", "Телефон уже существует.")
            return

        if student_id:
            self.cursor.execute('''
                UPDATE students SET first_name=?, last_name=?, middle_name=?, birth_date=?, phone=?, email=?, address=?, group_id=? WHERE id=?
            ''', (data['Имя'], data['Фамилия'], data['Отчество'], data['Дата рождения'], data['Телефон'], data['Email'],
                  data['Адрес'], group_id, student_id))
        else:
            self.cursor.execute('''
                INSERT INTO students (first_name, last_name, middle_name, birth_date, phone, email, address, group_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (data['Имя'], data['Фамилия'], data['Отчество'], data['Дата рождения'], data['Телефон'], data['Email'],
                  data['Адрес'], group_id))
        self.conn.commit()
        self.load_data()
        window.destroy()

    def save_event(self, entries, window, event_id=None, student_id=None):
        data = {label: entry.get() for label, entry in entries.items()}
        if event_id:
            self.cursor.execute('''
                UPDATE events SET date=?, title=?, description=?, category=? WHERE id=?
            ''', (data['Дата'], data['Название'], data['Описание'], data['Категория'], event_id))
        else:
            self.cursor.execute('''
                INSERT INTO events (student_id, date, title, description, category) VALUES (?, ?, ?, ?, ?)
            ''', (student_id, data['Дата'], data['Название'], data['Описание'], data['Категория']))
        self.conn.commit()
        self.load_data()
        window.destroy()

    def save_group(self, group_name, window, group_id=None):
        if not group_name.strip():
            messagebox.showerror("Ошибка валидации", "Пожалуйста, заполните поле 'Группа'.")
            return
        if group_id:
            self.cursor.execute('UPDATE groups SET group_name=? WHERE id=?', (group_name, group_id))
        else:
            self.cursor.execute('INSERT INTO groups (group_name) VALUES (?)', (group_name,))
        self.conn.commit()
        self.load_data()
        window.destroy()

    def check_unique_email(self, email, student_id=None):
        query = "SELECT id FROM students WHERE email = ?"
        params = [email]
        if student_id:
            query += " AND id != ?"
            params.append(student_id)
        self.cursor.execute(query, params)
        return self.cursor.fetchone() is None

    def check_unique_phone(self, phone, student_id=None):
        query = "SELECT id FROM students WHERE phone = ?"
        params = [phone]
        if student_id:
            query += " AND id != ?"
            params.append(student_id)
        self.cursor.execute(query, params)
        return self.cursor.fetchone() is None

    def delete_selected_item(self, treeview, query):
        selected_item = treeview.selection()
        if selected_item:
            values = treeview.item(selected_item)['values']
            self.cursor.execute(query, (values[0],))
            self.conn.commit()
            treeview.delete(selected_item)
        else:
            messagebox.showwarning("Предупреждение", "Пожалуйста, выберите элемент для удаления.")

    def edit_selected_item(self, treeview, open_window_func):
        selected_item = treeview.selection()
        if selected_item:
            values = treeview.item(selected_item)['values']
            open_window_func('Редактировать', values)
        else:
            messagebox.showwarning("Предупреждение", "Пожалуйста, выберите элемент для редактирования.")

    def show_student_context_menu(self, event):
        selected_item = self.tree_students.identify_row(event.y)
        if selected_item:
            self.tree_students.selection_set(selected_item)
            student = self.tree_students.item(selected_item)['values']
            context_menu = tk.Menu(self.root, tearoff=0)
            context_menu.add_command(label="Сведения о студенте",
                                     command=lambda: self.show_student_events(student[0], show_info=True))
            context_menu.post(event.x_root, event.y_root)

    def show_student_events(self, student_id, show_info=False):
        events_window = tk.Toplevel(self.root)
        events_window.title("События студента")
        tree_events = self.create_treeview(events_window, ['ID события', 'Дата', 'Название', 'Описание', 'Категория'])
        self.cursor.execute("SELECT id, date, title, description, category FROM events WHERE student_id=?",
                            (student_id,))
        events = self.cursor.fetchall()
        for event in events:
            tree_events.insert('', 'end', values=event)

        if show_info:
            self.cursor.execute('SELECT start_date, end_date, group_name FROM education_periods WHERE student_id=?',
                                (student_id,))
            education_periods = self.cursor.fetchall()
            info_text = ""
            for period in education_periods:
                start_date, end_date, group_name = period
                info_text += f"Дата начала обучения: {start_date}\n"
                info_text += f"Дата окончания обучения: {end_date}\n"
                info_text += f"Группа: {group_name}\n\n"
            info_label = tk.Label(events_window, text=info_text)
            info_label.pack(padx=10, pady=10)

        button_frame = tk.Frame(events_window)
        button_frame.pack(fill=tk.X, pady=10)
        tk.Button(button_frame, text="Добавить событие",
                  command=lambda: self.open_event_window('Добавить событие', None, student_id)).pack(side=tk.LEFT,
                                                                                                     padx=5)
        tk.Button(button_frame, text="Редактировать событие",
                  command=lambda: self.edit_selected_event(tree_events, student_id)).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Удалить событие", command=lambda: self.delete_event(tree_events)).pack(
            side=tk.LEFT, padx=5)
        tk.Button(events_window, text="Закрыть", command=events_window.destroy).pack(pady=10)

    def edit_selected_event(self, treeview, student_id):
        selected_item = treeview.selection()
        if selected_item:
            values = treeview.item(selected_item)['values']
            self.open_event_window('Редактировать событие', values, student_id)
        else:
            messagebox.showwarning("Предупреждение", "Пожалуйста, выберите событие для редактирования.")

    def delete_event(self, treeview):
        selected_item = treeview.selection()
        if selected_item:
            values = treeview.item(selected_item)['values']
            self.cursor.execute("DELETE FROM events WHERE id=?", (values[0],))
            self.conn.commit()
            treeview.delete(selected_item)
        else:
            messagebox.showwarning("Предупреждение", "Пожалуйста, выберите событие для удаления.")

    def search_in_treeview(self, query, data, treeview, column=None):
        results = []
        for item in data:
            if any(query in str(value).lower() for value in item) if column is None else query in str(
                    item[column]).lower():
                results.append(item)
        self.update_treeview(treeview, results)

    def generate_group_report(self):
        window = tk.Toplevel(self.root)
        window.title("Отчёт по группе за период")
        labels = ['Группа', 'Дата начала', 'Дата окончания']
        entries = {}
        tk.Label(window, text='Группа').grid(row=0, column=0, padx=10, pady=5)
        group_combobox = ttk.Combobox(window, values=[group[1] for group in self.groups])
        group_combobox.grid(row=0, column=1, padx=10, pady=5)
        entries['Группа'] = group_combobox
        for i, label_text in enumerate(labels[1:], start=1):
            tk.Label(window, text=label_text).grid(row=i, column=0, padx=10, pady=5)
            entry = DateEntry(window, date_pattern='dd.MM.yyyy', locale='ru')
            entry.grid(row=i, column=1, padx=10, pady=5)
            entries[label_text] = entry
        tk.Button(window, text="Сгенерировать отчёт", command=lambda: self.create_group_report(entries, window)).grid(
            row=len(labels), column=0, columnspan=2, pady=10)

    def create_group_report(self, entries, window):
        group_name = entries['Группа'].get()
        start_date = entries['Дата начала'].get()
        end_date = entries['Дата окончания'].get()

        if not group_name or not start_date or not end_date:
            messagebox.showwarning("Ошибка", "Пожалуйста, заполните все поля.")
            return

        try:
            self.cursor.execute('''
                SELECT s.id, s.first_name, s.last_name, s.middle_name, e.date, e.title, e.description, e.category
                FROM students s
                JOIN events e ON s.id = e.student_id
                JOIN groups g ON s.group_id = g.id
                WHERE g.group_name = ? AND e.date >= ? AND e.date <= ?
            ''', (group_name, start_date, end_date))
            report_data = self.cursor.fetchall()
        except sqlite3.Error as e:
            messagebox.showerror("Ошибка базы данных", f"Ошибка: {e}")
            return

        if not report_data:
            messagebox.showinfo("Информация", "Нет данных для указанного периода.")
            return

        report_window = tk.Toplevel(self.root)
        report_window.title(f"Отчёт по группе {group_name} за период с {start_date} по {end_date}")
        tree_report = self.create_treeview(report_window,
                                           ['ID студента', 'Фамилия', 'Имя', 'Отчество', 'Дата события', 'Название',
                                            'Описание', 'Категория'])
        for row in report_data:
            tree_report.insert('', 'end', values=row)
        tk.Button(report_window, text="Закрыть", command=report_window.destroy).pack(pady=10)
        window.destroy()

    def import_from_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            data = self.parse_excel(file_path)
            self.import_data(data)

    def import_data(self, data):
        for student in data:
            self.cursor.execute('''
                INSERT INTO students (first_name, last_name, middle_name, birth_date, phone, email, address, group_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (student['Имя'], student['Фамилия'], student['Отчество'], student['Дата рождения'],
                  student['Телефон'], student['Email'], student['Адрес'], self.get_group_id(student['Группа'])))
        self.conn.commit()
        self.load_data()

    def parse_excel(self, file_path):
        workbook = load_workbook(filename=file_path)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        return [{headers[i]: cell.value for i, cell in enumerate(row)} for row in worksheet.iter_rows(min_row=2)]

    def export_to_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.write_excel(file_path, self.students)

    def write_excel(self, file_path, data):
        workbook = Workbook()
        worksheet = workbook.active
        headers = data[0].keys() if isinstance(data[0], dict) else ['ID', 'Фамилия', 'Имя', 'Отчество', 'Дата рождения',
                                                                    'Телефон', 'Email', 'Адрес', 'Группа']
        worksheet.append(headers)
        for row in data:
            worksheet.append(row.values() if isinstance(row, dict) else row)
        workbook.save(file_path)

    def get_group_id(self, group_name):
        self.cursor.execute('SELECT id FROM groups WHERE group_name = ?', (group_name,))
        group = self.cursor.fetchone()
        return group[0] if group else None

    def get_student_by_id(self, student_id):
        self.cursor.execute("SELECT * FROM students WHERE id=?", (student_id,))
        return self.cursor.fetchone()

    def setup_api(self):
        app = Flask(__name__)

        @app.route('/api/students', methods=['GET'])
        def get_students():
            self.cursor.execute("SELECT * FROM students")
            return jsonify(self.cursor.fetchall())

        @app.route('/api/events', methods=['GET'])
        def get_events():
            self.cursor.execute("SELECT * FROM events")
            return jsonify(self.cursor.fetchall())

        @app.route('/api/students/<int:student_id>/events', methods=['GET'])
        def get_student_events(student_id):
            self.cursor.execute("SELECT * FROM events WHERE student_id=?", (student_id,))
            return jsonify(self.cursor.fetchall())

        @app.route('/api/groups', methods=['GET'])
        def get_groups():
            self.cursor.execute("SELECT * FROM groups")
            return jsonify(self.cursor.fetchall())

        self.api_app = app
        self.api_thread = Thread(target=self.api_app.run, kwargs={'port': 5000})
        self.api_thread.start()


if __name__ == "__main__":
    root = tk.Tk()
    app = StudentManagementSystem(root)
    root.mainloop()